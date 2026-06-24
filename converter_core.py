#!/usr/bin/env python3
"""
converter_core.py
=================
Pure-Python VIKING / Color Line Cargo invoice -> Excel converter.

Unlike pdfToExcel.py (which uses pdfplumber), this module reads PDFs with
pdfminer.six only, so it has NO binary dependencies and can run in the
browser through Pyodide. The parsing rules and Excel output are identical.

Public entry point:
    xlsx_bytes = convert_pdf_bytes(pdf_bytes, cfg=Config())
"""

from __future__ import annotations

import datetime as dt
import io
import re
from dataclasses import dataclass
from typing import Optional

from pdfminer.high_level import extract_pages
from pdfminer.layout import LAParams, LTTextContainer, LTTextLine
from openpyxl import Workbook
from openpyxl.styles import Font


# ============================================================================
# 1. CONFIGURATION  — same knobs as the desktop pdfToExcel.py.
#    (Folder/CLI options from the desktop version are omitted here: the browser
#     has no file system, files come straight from the page.)
# ============================================================================
@dataclass
class Config:
    pdf_glob: str = "*.pdf"            # which files to pick up (used by the page)

    # --- parsing behaviour -------------------------------------------------
    # Include the "With Following Configuration:" block in the description?
    include_config_block: bool = True

    # How to fill the Kjøper / Mottaker (buyer / recipient) columns:
    #   "standard"          -> top-left block = buyer, delivery block = recipient
    #                          (reproduces sample file 90920196 exactly)
    #   "delivery_as_buyer" -> delivery block = buyer, recipient left blank
    #                          (reproduces sample file 90920200)
    address_mode: str = "standard"

    # --- fixed / default values that aren't reliably in the PDF ------------
    transport_mode_default: str = "Postal"
    seller_country_default: str = "Denmark"

    # --- geometry / formatting --------------------------------------------
    col_split_x: float = 300.0        # x that divides the two address columns
    addr_y_max: float = 185.0         # only read addresses above this y
    thousands_sep: str = "\u00a0"     # non-breaking space, as in the samples
    decimal_sep: str = ","


# ============================================================================
# 2. COLUMN DEFINITIONS  — reorder/rename here and the writer follows.
# ============================================================================
HOVED_COLS = [
    "Faktura", "Fakturadato", "Totalbeløp", "Leveringsbetingelse", "Leveringssted",
    "Transportmåte", "Transportør", "Bruttovekt (kg)", "Nettovekt (kg)", "Varelinjer",
    "Selger navn", "Selger gateadresse", "Selger postnummer", "Selger poststed",
    "Selger land", "Kjøper navn", "Kjøper gateadresse", "Kjøper postnummer",
    "Kjøper poststed", "Kjøper land", "Avsender navn", "Avsender gateadresse",
    "Avsender postnummer", "Avsender poststed", "Avsender land", "Mottaker navn",
    "Mottaker gateadresse", "Mottaker postnummer", "Mottaker poststed", "Mottaker land",
]
VARE_COLS = [
    "Linjenummer", "Beskrivelse", "Tariffkode", "Varekode", "Opprinnelsesland",
    "Antall", "Salgsenhet", "Salgspris per enhet", "Linjebeløp", "Valuta",
    "Rabatt %", "Gebyr %", "Bruttovekt (kg)", "Nettovekt (kg)",
]

# lines that must never be folded into a line-item description
EXCLUDE_PREFIXES = (
    "Price", "H.S. Code", "Country of origin", "Serial Number", "Order",
    "Delivery note", "Net Value for Item", "SD IC Discount", "Approval status",
    "Labor time", "Transferred", "After due date", "Amount ex", "Delivery :",
    "Packing", "Weight :", "Volume", "The exporter", "document customs",
    "products are", "where otherwise", "Head Office", "VIKING LIFE-SAVING",
    "Saedding", "DK-6710", "IBAN", "Item Material", "Terms of delivery",
    "Your reference", "Multiple", "Terms of payment", "Current month",
    "PROFORMA", "INTERCOMPANY", "DAP",
    # The four below are pdfminer-only: it keeps the footer's right-hand column
    # as separate lines, so these would otherwise leak into a description.
    "Tel :", "Fax :", "Web :", "E-mail :",
)
ITEM_RE = re.compile(
    r"^(\d+)\s+(\S+)\s+(.+?)\s+([\d.,]+)\s+([A-Z]{2,4})(?:\s+([\d.,]+))?$"
)
CONFIG_HEADER = "With Following Configuration:"


# ============================================================================
# PDF READER  (pdfminer.six only)
# ============================================================================
class Page:
    """One PDF page reduced to positioned text fragments."""

    def __init__(self, layout):
        self.height = layout.height
        self.frags: list[tuple[int, float, str]] = []   # (top, x0, text)
        self._collect(layout)

    def _collect(self, obj):
        if isinstance(obj, LTTextLine):
            txt = obj.get_text().rstrip("\n")
            if txt.strip():
                top = round(self.height - obj.bbox[3])    # bottom-up -> top-down
                self.frags.append((top, obj.bbox[0], txt))
        elif isinstance(obj, LTTextContainer):
            for child in obj:
                self._collect(child)
        elif hasattr(obj, "__iter__"):
            for child in obj:
                self._collect(child)

    @staticmethod
    def _group(frags, tol=2) -> list[str]:
        """Merge fragments that share a baseline into one left-to-right line."""
        frags = sorted(frags, key=lambda f: (f[0], f[1]))
        lines, cur, cy = [], [], None
        for top, x0, txt in frags:
            if cy is None or abs(top - cy) <= tol:
                cur.append((x0, txt))
                cy = top if cy is None else cy
            else:
                lines.append(" ".join(t for _, t in sorted(cur)))
                cur, cy = [(x0, txt)], top
        if cur:
            lines.append(" ".join(t for _, t in sorted(cur)))
        return lines

    def merged_lines(self) -> list[str]:
        """Full-width reading order (≈ pdfplumber extract_text)."""
        return self._group(self.frags)

    def band_lines(self, x_lo, x_hi, y_max=None) -> list[str]:
        """Lines whose fragments start within [x_lo, x_hi) (one column)."""
        sub = [f for f in self.frags
               if x_lo <= f[1] < x_hi and (y_max is None or f[0] < y_max)]
        return self._group(sub)


class PdfDoc:
    def __init__(self, source, cfg: Config):
        pages_layout = list(extract_pages(source, laparams=LAParams()))
        self.pages = [Page(p) for p in pages_layout]
        self.cfg = cfg

    @property
    def flat_lines(self) -> list[str]:
        out = []
        for p in self.pages:
            out.extend(p.merged_lines())
        return out

    def first_page_columns(self):
        p = self.pages[0]
        left = p.band_lines(0, self.cfg.col_split_x, self.cfg.addr_y_max)
        right = p.band_lines(self.cfg.col_split_x, 10_000, self.cfg.addr_y_max)
        return left, right

    def footer_left_lines(self) -> list[str]:
        """Left column of the footer (contains the seller address block)."""
        out = []
        for p in self.pages:
            out.extend(p.band_lines(0, 250, y_max=None))
        return out


# ============================================================================
# SMALL UTILITIES  (identical to desktop version)
# ============================================================================
def euro_num(s: Optional[str]) -> Optional[float]:
    if s is None:
        return None
    s = s.strip().replace("\u00a0", " ")
    m = re.search(r"-?\d[\d.\s]*(?:,\d+)?", s)
    if not m:
        return None
    t = m.group(0).replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return None


def fmt_amount(value, currency, cfg: Config) -> str:
    if value is None:
        return ""
    whole = f"{value:,.2f}".replace(",", cfg.thousands_sep).replace(".", cfg.decimal_sep)
    return f"{whole} {currency}".strip()


def titlecase_place(s: str) -> str:
    return " ".join(
        w.capitalize() if not re.match(r"^[A-Z]{1,2}-?\d", w) else w
        for w in s.split()
    )


# Characters that spreadsheet apps may treat as the start of a formula.
# A cell beginning with one of these (taken straight from a PDF) could run a
# formula when the resulting .xlsx is opened. Prefixing it with an apostrophe
# forces the cell to be plain text. Real invoice fields never start with these,
# so normal output is unchanged — this only ever touches hostile/odd content.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def sanitize_cell(v):
    """Neutralise spreadsheet-formula injection in text cells."""
    if isinstance(v, str) and v[:1] in _FORMULA_LEAD:
        return "'" + v
    return v


def parse_address(lines, drop_labels):
    cleaned = [
        l.strip() for l in lines
        if l.strip()
        and not any(l.strip().startswith(d) for d in drop_labels)
        and not l.strip().upper().startswith("ATT.")
    ]
    postal_re = re.compile(r"^([A-Z]{0,2}-?\d{3,5})\s+(.+)$")
    name_lines, street, postal, place, country = [], "", "", "", ""
    for l in cleaned:
        m = postal_re.match(l)
        if m and not place:
            postal, place = m.group(1), titlecase_place(m.group(2))
            continue
        if l.upper() in ("NORWAY", "DENMARK", "SWEDEN", "GERMANY", "FRANCE"):
            country = l.capitalize()
            continue
        if not street and re.search(r"\d", l):
            street = l
            continue
        if not street:
            name_lines.append(l)
    return " ".join(name_lines), street, postal, place, country


# ============================================================================
# CORE PARSE
# ============================================================================
def parse_invoice(doc: PdfDoc, cfg: Config):
    flat = [re.sub(r"\s+", " ", l).strip() for l in doc.flat_lines]
    all_text = "\n".join(flat)
    left_lines, right_lines = doc.first_page_columns()

    mdoc = re.search(
        r"(PROFORMA INVOICE|INTERCOMPANY BILLING / TAX INVOICE)\s+(\d+)", all_text)
    invoice_no = mdoc.group(2) if mdoc else ""

    mdate = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", all_text)
    inv_date = (dt.date(int(mdate.group(3)), int(mdate.group(2)), int(mdate.group(1)))
                if mdate else None)

    mcur = re.search(r"Amount in (\w{3})", all_text)
    currency = mcur.group(1) if mcur else "DKK"

    terms = ""
    for l in flat:
        m = re.match(r"^\s*(DAP|EXW|FCA|CIF|FOB|DDP|CPT|DAT)\b", l)
        if m:
            terms = m.group(1)
            break

    delivery_place = gross_weight = total_value = None
    for l in flat:
        m = re.search(r"Delivery\s*:\s*(\S+)\s*(.*)$", l)
        if m:
            delivery_place = m.group(2).strip()
        m = re.search(r"Weight\s*:\s*([\d.,]+)\s*kg", l)
        if m:
            gross_weight = euro_num(m.group(1))
            if gross_weight is not None:
                gross_weight = round(gross_weight, 3)
    mtot = re.search(r"Amount ex\. VAT.*?\n\s*([\d.,]+)", all_text)
    if mtot:
        total_value = euro_num(mtot.group(1))

    topleft = parse_address(left_lines, ["Delivery address", "Company"])
    delivery = parse_address(right_lines, ["Delivery address", "Company"])
    if cfg.address_mode == "delivery_as_buyer":
        buyer, recipient = delivery, ("", "", "", "", "")
    else:
        buyer, recipient = topleft, delivery

    seller = parse_seller(doc.footer_left_lines(), cfg)

    items = parse_items(flat, currency, cfg)
    header = {
        "invoice_no": invoice_no, "date": inv_date, "currency": currency,
        "total": total_value, "terms": terms, "delivery_place": delivery_place or "",
        "gross_weight": gross_weight, "buyer": buyer, "recipient": recipient,
        "seller": seller, "n_items": len(items),
    }
    return header, items


def parse_seller(footer_lines, cfg: Config):
    name = "VIKING LIFE-SAVING EQUIPMENT A/S"
    street = postal = place = ""
    for i, l in enumerate(footer_lines):
        if l.strip() == name:
            rest = [x.strip() for x in footer_lines[i + 1:i + 6] if x.strip()]
            for r in rest:
                m = re.match(r"^([A-Z]{0,2}-?\d{3,5})\s+(.+)$", r)
                if m and not place:
                    postal, place = m.group(1), titlecase_place(m.group(2))
                elif not street and re.search(r"\d", r) and "Acct" not in r:
                    street = r
            break
    return (name, street, postal, place, cfg.seller_country_default)


def parse_items(flat, currency, cfg: Config):
    starts = [i for i, l in enumerate(flat) if ITEM_RE.match(l)]
    items = []
    for idx, s in enumerate(starts):
        end = starts[idx + 1] if idx + 1 < len(starts) else len(flat)
        m = ITEM_RE.match(flat[s])
        material = m.group(2)
        desc_parts = [m.group(3).strip()]
        qty = euro_num(m.group(4))
        unit = m.group(5)
        unit_price = euro_num(m.group(6)) if m.group(6) else None

        tariff = origin = ""
        net_value = price_value = discount = None
        in_config = False
        for l in flat[s + 1:end]:
            t = l.strip()
            if not t:
                continue
            mm = re.match(r"H\.S\. Code:\s*(\d+)", t)
            if mm:
                tariff = mm.group(1); continue
            mm = re.match(r"Country of origin:\s*([A-Z]{2})", t)
            if mm:
                origin = mm.group(1); continue
            mm = re.match(r"Net Value for Item\s+([\d.,]+)", t)
            if mm:
                net_value = euro_num(mm.group(1)); continue
            mm = re.match(r"Price\s+([\d.,]+)", t)
            if mm:
                price_value = euro_num(mm.group(1)); continue
            mm = re.search(r"SD IC Discount %\s+([\d.,]+)\s*%", t)
            if mm:
                discount = euro_num(mm.group(1)); continue
            if t.startswith(CONFIG_HEADER):
                in_config = True
                if cfg.include_config_block:
                    desc_parts.append(t)
                continue
            if in_config and not cfg.include_config_block:
                continue
            if any(t.startswith(p) for p in EXCLUDE_PREFIXES):
                continue
            if re.search(r"declares that", t):
                continue
            if re.match(r"^[\d.,]+(?:\s+[\d.,]+)+$", t):
                continue
            if re.match(r"^[\d.,]+$", t):
                continue
            desc_parts.append(t)

        line_amount = net_value if net_value is not None else price_value
        if line_amount is None and unit_price is not None and qty:
            line_amount = round(unit_price * qty, 2)
        if unit_price is None and line_amount is not None and qty:
            unit_price = round(line_amount / qty, 2)

        items.append({
            "line": len(items) + 1,
            "description": " ".join(p for p in desc_parts if p).strip(),
            "tariff": tariff, "material": material, "origin": origin,
            "qty": round(qty, 3) if qty is not None else None, "unit": unit,
            "unit_price": round(unit_price, 2) if unit_price is not None else None,
            "line_amount": round(line_amount, 2) if line_amount is not None else None,
            "currency": currency,
            "discount": round(discount, 2) if discount is not None else None,
        })
    return items


# ============================================================================
# WRITE WORKBOOK  ->  bytes
# ============================================================================
def build_workbook_bytes(header, items, cfg: Config) -> bytes:
    wb = Workbook()
    hov = wb.active
    hov.title = "Hoveddata"
    bold = Font(bold=True)

    hov.append(HOVED_COLS)
    s = header["seller"]; b = header["buyer"]; r = header["recipient"]
    hov.append([sanitize_cell(v) for v in [
        int(header["invoice_no"]) if header["invoice_no"] else "",
        header["date"], fmt_amount(header["total"], header["currency"], cfg),
        header["terms"], header["delivery_place"], cfg.transport_mode_default, "",
        header["gross_weight"], "", header["n_items"],
        s[0], s[1], s[2], s[3], s[4], b[0], b[1], b[2], b[3], b[4],
        "", "", "", "", "", r[0], r[1], r[2], r[3], r[4],
    ]])
    if header["date"]:
        hov.cell(row=2, column=2).number_format = "yyyy-mm-dd"

    var = wb.create_sheet("Varelinjer")
    var.append(VARE_COLS)
    for it in items:
        var.append([sanitize_cell(v) for v in [
            it["line"], it["description"], it["tariff"], it["material"],
            it["origin"], it["qty"], it["unit"], it["unit_price"],
            it["line_amount"], it["currency"], it["discount"], "", "", "",
        ]])

    for ws, cols in ((hov, HOVED_COLS), (var, VARE_COLS)):
        for c in range(1, len(cols) + 1):
            ws.cell(row=1, column=c).font = bold
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None),
                        default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def convert_pdf_bytes(pdf_bytes: bytes, cfg: Config = None) -> bytes:
    """Main entry point: PDF bytes in, .xlsx bytes out."""
    cfg = cfg or Config()
    doc = PdfDoc(io.BytesIO(pdf_bytes), cfg)
    header, items = parse_invoice(doc, cfg)
    return build_workbook_bytes(header, items, cfg)
